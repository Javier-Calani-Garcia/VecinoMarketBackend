"""CU18/CU19: exportación de reportes de dashboard a CSV, XLSX y PDF.

Formato común de entrada para no repetir la generación de cada tipo de
archivo por cada dashboard: una lista de "secciones", cada una con un
título, encabezados de columna y filas de datos ya formateados a texto.
"""
import csv
import io
from datetime import datetime

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

COLOR_MARCA = '#D97706'
FORMATOS_VALIDOS = ('csv', 'xlsx', 'pdf')


def _nombre_archivo(base, extension):
    fecha = datetime.now().strftime('%Y%m%d_%H%M')
    return f'{base}_{fecha}.{extension}'


def exportar_reporte(formato, base_nombre, titulo, subtitulo, secciones):
    """secciones: [{'titulo': str, 'headers': [...], 'filas': [[...], ...]}]"""
    if formato == 'csv':
        return _csv_response(base_nombre, titulo, subtitulo, secciones)
    if formato == 'xlsx':
        return _xlsx_response(base_nombre, titulo, subtitulo, secciones)
    if formato == 'pdf':
        return _pdf_response(base_nombre, titulo, subtitulo, secciones)
    raise ValueError(f'Formato de exportación no soportado: {formato}')


def _csv_response(base_nombre, titulo, subtitulo, secciones):
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow([titulo])
    writer.writerow([subtitulo])
    for seccion in secciones:
        writer.writerow([])
        writer.writerow([seccion['titulo']])
        writer.writerow(seccion['headers'])
        for fila in seccion['filas']:
            writer.writerow(fila)

    contenido = '﻿' + buffer.getvalue()  # BOM para que Excel abra bien los acentos
    response = HttpResponse(contenido, content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="{_nombre_archivo(base_nombre, "csv")}"'
    return response


def _xlsx_response(base_nombre, titulo, subtitulo, secciones):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Reporte'

    fila_actual = 1
    ws.cell(row=fila_actual, column=1, value=titulo).font = Font(size=14, bold=True, color='D97706')
    fila_actual += 1
    ws.cell(row=fila_actual, column=1, value=subtitulo).font = Font(italic=True, color='666666')
    fila_actual += 2

    for seccion in secciones:
        ws.cell(row=fila_actual, column=1, value=seccion['titulo']).font = Font(bold=True, size=12)
        fila_actual += 1

        for col, encabezado in enumerate(seccion['headers'], start=1):
            celda = ws.cell(row=fila_actual, column=col, value=encabezado)
            celda.font = Font(bold=True, color='FFFFFF')
            celda.fill = PatternFill('solid', fgColor=COLOR_MARCA.lstrip('#'))
            celda.alignment = Alignment(horizontal='left')
        fila_actual += 1

        filas = seccion['filas'] or [['Sin datos']]
        for fila in filas:
            for col, valor in enumerate(fila, start=1):
                ws.cell(row=fila_actual, column=col, value=valor)
            fila_actual += 1
        fila_actual += 1  # espacio entre secciones

    for col in range(1, 6):
        ws.column_dimensions[get_column_letter(col)].width = 26

    buffer = io.BytesIO()
    wb.save(buffer)
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{_nombre_archivo(base_nombre, "xlsx")}"'
    return response


def _pdf_response(base_nombre, titulo, subtitulo, secciones):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=letter,
        topMargin=1.5 * cm, bottomMargin=1.5 * cm, leftMargin=1.5 * cm, rightMargin=1.5 * cm,
    )
    estilos = getSampleStyleSheet()
    estilo_titulo = ParagraphStyle('TituloReporte', parent=estilos['Heading1'], textColor=colors.HexColor(COLOR_MARCA))
    estilo_subtitulo = ParagraphStyle('Subtitulo', parent=estilos['Normal'], textColor=colors.HexColor('#666666'), spaceAfter=14)
    estilo_seccion = ParagraphStyle('Seccion', parent=estilos['Heading2'], spaceBefore=16, spaceAfter=6, fontSize=13)

    elementos = [Paragraph(titulo, estilo_titulo), Paragraph(subtitulo, estilo_subtitulo)]

    for seccion in secciones:
        elementos.append(Paragraph(seccion['titulo'], estilo_seccion))
        filas_texto = [[str(v) for v in fila] for fila in seccion['filas']]
        datos_tabla = [seccion['headers']] + (filas_texto or [['Sin datos'] + [''] * (len(seccion['headers']) - 1)])
        tabla = Table(datos_tabla, hAlign='LEFT')
        tabla.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(COLOR_MARCA)),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8.5),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#DDDDDD')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#FAFAFA')]),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        elementos.append(tabla)
        elementos.append(Spacer(1, 4))

    doc.build(elementos)
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{_nombre_archivo(base_nombre, "pdf")}"'
    return response
